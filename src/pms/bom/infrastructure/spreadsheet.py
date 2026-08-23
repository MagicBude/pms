"""不执行宏、公式或外部链接的 Office Open XML BOM 解析器。"""

from collections.abc import Mapping
from io import BytesIO
from pathlib import PurePath
from zipfile import BadZipFile, ZipFile, is_zipfile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from pms.bom.application.service import (
    MAX_BOM_SIZE_BYTES,
    BomImportError,
    ParsedSpreadsheetRow,
)

ALLOWED_EXTENSIONS = frozenset({".xlsx", ".xlsm"})
REQUIRED_FIELDS = frozenset({"material_name", "quantity_per_unit", "unit"})
ALLOWED_FIELDS = frozenset(
    {
        "material_code",
        "material_name",
        "specification",
        "brand",
        "quantity_per_unit",
        "unit",
        "level_path",
        "assembly_code",
        "assembly_name",
        "remark",
    }
)
MAX_EXPANDED_WORKBOOK_BYTES = 100 * 1024 * 1024


class OpenPyxlBomSpreadsheetParser:
    """只读取第一个工作表的映射列，不计算任何工作簿逻辑。

    ``openpyxl`` 以 ``read_only`` 和 ``data_only=False`` 打开 OOXML；后者
    保留公式标记，使系统能够拒绝公式而不是误用缓存值。``keep_links``
    关闭外部链接保留，`.xlsm` 的 VBA 二进制也不会加载或执行。
    """

    def parse(
        self, *, filename: str, content: bytes, mapping: Mapping[str, str]
    ) -> list[ParsedSpreadsheetRow]:
        extension = PurePath(filename).suffix.lower()
        if extension not in ALLOWED_EXTENSIONS:
            raise BomImportError("BOM 文件只支持 .xlsx 或 .xlsm。")
        if not content or len(content) > MAX_BOM_SIZE_BYTES:
            raise BomImportError("BOM 文件必须大于 0 byte 且不超过 25 MiB。")
        if not is_zipfile(BytesIO(content)):
            raise BomImportError("文件内容不是有效的 Office Open XML 工作簿。")
        self._validate_expanded_size(content)
        normalized_mapping = self._validate_mapping(mapping)
        try:
            workbook = load_workbook(
                BytesIO(content),
                read_only=True,
                data_only=False,
                keep_links=False,
                keep_vba=False,
            )
        except (BadZipFile, InvalidFileException, KeyError, OSError, ValueError) as error:
            raise BomImportError("无法安全读取 BOM 工作簿。") from error
        try:
            worksheet = workbook.worksheets[0]
            rows = worksheet.iter_rows()
            try:
                header_cells = next(rows)
            except StopIteration as error:
                raise BomImportError("BOM 工作表为空。") from error
            headers = {
                str(cell.value).strip(): index
                for index, cell in enumerate(header_cells)
                if cell.value is not None and str(cell.value).strip()
            }
            missing_headers = sorted(set(normalized_mapping.values()) - headers.keys())
            if missing_headers:
                raise BomImportError(f"字段映射引用了不存在的表头：{', '.join(missing_headers)}。")
            parsed: list[ParsedSpreadsheetRow] = []
            for source_row_number, cells in enumerate(rows, start=2):
                values: dict[str, str] = {}
                formula_fields: list[str] = []
                for field, header in normalized_mapping.items():
                    index = headers[header]
                    cell = cells[index] if index < len(cells) else None
                    if cell is not None and cell.data_type == "f":
                        formula_fields.append(field)
                    raw_value = None if cell is None else cell.value
                    values[field] = "" if raw_value is None else str(raw_value).strip()
                if any(values.values()) or formula_fields:
                    parsed.append(
                        ParsedSpreadsheetRow(
                            source_row_number=source_row_number,
                            values=values,
                            formula_fields=tuple(formula_fields),
                        )
                    )
            if not parsed:
                raise BomImportError("BOM 工作表没有可导入的数据行。")
            return parsed
        finally:
            workbook.close()

    @staticmethod
    def _validate_mapping(mapping: object) -> dict[str, str]:
        if not isinstance(mapping, dict):
            raise BomImportError("字段映射格式无效。")
        normalized: dict[str, str] = {}
        for raw_field, raw_header in mapping.items():
            field = str(raw_field).strip()
            header = str(raw_header).strip()
            if field not in ALLOWED_FIELDS or not header:
                raise BomImportError("字段映射包含未知字段或空表头。")
            normalized[field] = header
        missing_fields = sorted(REQUIRED_FIELDS - normalized.keys())
        if missing_fields:
            raise BomImportError(f"字段映射缺少必填字段：{', '.join(missing_fields)}。")
        if len(set(normalized.values())) != len(normalized):
            raise BomImportError("一个工作表列不能同时映射为多个 BOM 字段。")
        return normalized

    @staticmethod
    def _validate_expanded_size(content: bytes) -> None:
        """拒绝压缩后很小、解压后异常巨大的 ZIP，降低资源耗尽风险。"""
        try:
            with ZipFile(BytesIO(content)) as archive:
                expanded_size = sum(entry.file_size for entry in archive.infolist())
        except (BadZipFile, OSError) as error:
            raise BomImportError("文件内容不是有效的 Office Open XML 工作簿。") from error
        if expanded_size > MAX_EXPANDED_WORKBOOK_BYTES:
            raise BomImportError("BOM 工作簿解压后超过安全上限。")
