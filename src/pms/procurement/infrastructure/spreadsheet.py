"""把不可变订单快照渲染成不包含宏的稳定 Excel 单据。"""

from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal
from io import BytesIO
from typing import cast

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet


@dataclass(frozen=True, slots=True)
class OrderExportLine:
    project_code: str
    request_number: str
    material_code: str
    material_name: str
    part_attribute: str
    unit: str
    quantity: Decimal
    unit_price: Decimal
    tax_rate: Decimal
    net_amount: Decimal
    tax_amount: Decimal
    gross_amount: Decimal
    remark: str


@dataclass(frozen=True, slots=True)
class OrderExportData:
    order_number: str
    status: str
    kind: str
    supplier_name: str
    currency: str
    issued_at: datetime
    lines: tuple[OrderExportLine, ...]
    net_amount: Decimal
    tax_amount: Decimal
    gross_amount: Decimal


def render_order_xlsx(data: OrderExportData) -> bytes:
    """生成只含冻结字段的 xlsx；不读取可变主数据或执行公式。"""
    workbook = Workbook()
    sheet = cast(Worksheet, workbook.active)
    sheet.title = "正式订单"
    sheet.merge_cells("A1:M1")
    sheet["A1"] = "采购 / 外协订单"
    sheet["A1"].font = Font(size=16, bold=True)
    sheet["A1"].alignment = Alignment(horizontal="center")
    sheet.append(["订单号", data.order_number, "类型", data.kind, "供应商", data.supplier_name])
    sheet.append(["币种", data.currency, "签发时间", data.issued_at.isoformat()])
    sheet.append([])
    headers = [
        "项目",
        "请购号",
        "物料编码",
        "物料名称",
        "零件属性",
        "单位",
        "数量",
        "单价",
        "税率(%)",
        "未税金额",
        "税额",
        "含税金额",
        "备注",
    ]
    sheet.append(headers)
    for cell in sheet[sheet.max_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
    for line in data.lines:
        sheet.append(
            [
                line.project_code,
                line.request_number,
                line.material_code,
                line.material_name,
                line.part_attribute,
                line.unit,
                line.quantity,
                line.unit_price,
                line.tax_rate,
                line.net_amount,
                line.tax_amount,
                line.gross_amount,
                line.remark,
            ]
        )
    sheet.append([])
    sheet.append(
        [
            "合计",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            data.net_amount,
            data.tax_amount,
            data.gross_amount,
        ]
    )
    sheet.freeze_panes = "A6"
    sheet.auto_filter.ref = f"A5:M{4 + len(data.lines)}"
    widths = (16, 20, 18, 28, 12, 10, 14, 14, 12, 16, 16, 16, 30)
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
